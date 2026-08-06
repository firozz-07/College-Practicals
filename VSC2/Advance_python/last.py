from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.drawing.image import Image
from openpyxl.chart import BarChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo

wb = Workbook()
ws = wb.active
ws.title = "Student Data"

data = [
    ["Name", "Math", "Science", "English"],
    ["Rahul", 80, 70, 90],
    ["Priya", 90, 85, 88],
    ["Amit", 75, 60, 70],
    ["Neha", 95, 98, 96]
]

for row in data:
    ws.append(row)

# Freeze
ws.freeze_panes = "B2"

# Filter
tab = Table(displayName="StudentTable", ref="A1:D5")

style = TableStyleInfo(
    name="TableStyleMedium9",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False
)

tab.tableStyleInfo = style
ws.add_table(tab)

# Formula
ws["E1"] = "Total"

for i in range(2, 6):
    ws[f"E{i}"] = f"=SUM(B{i}:D{i})"

# Style
fill = PatternFill(
    start_color="FFFF00",
    end_color="FFFF00",
    fill_type="solid"
)

font = Font(
    bold=True,
    color="FF0000",
    size=14
)

thin = Side(
    border_style="thin",
    color="000000"
)

for cell in ws[1]:
    cell.fill = fill
    cell.font = font
    cell.alignment = Alignment(horizontal="center")
    cell.border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

# Image (Place logo.png in same folder)
try:
    img = Image("logo.png")
    img.width = 120
    img.height = 120
    ws.add_image(img, "G2")
except:
    print("Image not found.")

# Bar Chart
chart = BarChart()

values = Reference(
    ws,
    min_col=2,
    max_col=4,
    min_row=1,
    max_row=5
)

cats = Reference(
    ws,
    min_col=1,
    min_row=2,
    max_row=5
)

chart.add_data(values, titles_from_data=True)
chart.set_categories(cats)
chart.title = "Student Marks"

ws.add_chart(chart, "G10")

# Save
wb.save("Student_Report.xlsx")

print("Excel file created successfully!")